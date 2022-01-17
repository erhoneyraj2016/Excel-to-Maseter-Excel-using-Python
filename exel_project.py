from re import X
from openpyxl import workbook, load_workbook

wb_data_sheet= load_workbook('input_data/input_data_sheet.xlsx')
wb_master_sheet= load_workbook('master_data/master_sheet.xlsx')

data_sheet_read=wb_data_sheet['INPUT-DATA']
master_sheet_write=wb_master_sheet['master_sheet']

ws = wb_data_sheet.active
ws1 = wb_master_sheet.active
#print(ws['A1'].value)


"""CREATING DICTIONARY FOR INPUT DATA"""
lst1=[]
count=0
dic={}
for row in data_sheet_read.rows:
    lst=[]
    for col in data_sheet_read.columns:
        lst.append(col[count].value)
    count+=1
    dic[row[0].value]=lst
"""----------------------------------"""


""" A LIST OF MASTER_SHEET-NAMES OF THE PRODUCTS"""
write=[]
for row1 in master_sheet_write.rows:
    write.append(row1[0].value)
"""--------------------------------------------"""

""" IF PRODUCT NAME IS NOT PRESENT IN THE MASTER FILE THEN IT WILL GET ADDED TO THE MASTER FILE"""
for key,value in dic.items():
    if key not in write:
        ws1.append(value)
    else:
        pass
wb_master_sheet.save("master_data/master_sheet.xlsx")
print('PRODUCT NAMES ADDED')
"""--------------------------------------------------------------------------------------------"""